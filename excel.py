import openpyxl
import re

# 打开工作簿
workbook = openpyxl.load_workbook('24PCM-PIN.xlsx')
workbook0 = openpyxl.load_workbook('CSIS-IAC 2023 PC member list V7 final.xueshanV1.xlsx')

# 获取工作表
sheet = workbook['Sheet1']
sheet0 = workbook0['Sheet2']
# 读取单元格数据
n=0
for v in sheet.values:
    n=n+1
    for A,B,C,D,E,F,G,H,I in sheet0.values:
        #stringv=str(v)
        #stringB=str(B)
        #判断特征
        
        #if stringB in stringv:
            #sheet.cell(n,9,str(D))
            #sheet.append({'A' : str(v), 'Q' : str(D)})
            #print(stringB)
            rv=re.split('[,]',str(v))
            r_namev=re.sub("[(]|[']|[ ]|[\"]","",rv[0])
            r_countryv=re.sub("[']|[\"]|[ ]","",rv[-2])
            rB=re.split('[,]',str(B))
            r_nameB=re.sub("[(]|[']|[ ]|[\"]","",rB[0])
            r_countryB=re.sub("[']|[\"]|[ ]","",rB[-1])
            if r_namev==r_nameB and r_countryv=="China":
                sheet.cell(n,9,str(D))
            if r_nameB=="HaoyaoChen" and r_namev==r_nameB:
                 print("XXXXXXXXXXXXXXXXXXXXXXXXXXXXX")

workbook.save('vc.xlsx')